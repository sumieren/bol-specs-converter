import csv, sys
import pandas # type: ignore
from openpyxl import load_workbook # type: ignore
from product import Product
from countryorders import CountryOrders


def main():
    # check if a single file was input
    if len(sys.argv) == 1:
        raise Exception("no file provided")
    if len(sys.argv) > 2:
        raise Exception("too many arguments provided")

    data = None
    file_name = sys.argv[1].split(".")[0].split("/")[1]

    # import the csv
    try:
        # File operations within the with block
        with open(sys.argv[1], "r") as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            data = list(csv_reader)
    except FileNotFoundError:
        print(f"Error: Could not find file '{sys.argv[1]}'")
        return
    except PermissionError:
        print(f"Error: No permission to read '{sys.argv[1]}'")
        return
    except Exception as e:
        print(f"Exception occurred: {e}. Did you save the CSV in unicode?")
        return

    # turns the read data into a list we can use
    header = data[0]
    rows = data[1:]

    nl_orders = CountryOrders("NL")
    be_orders = CountryOrders("BE")

    convert_data(data, nl_orders, be_orders)

    print(file_name)

    #nl_orders.report()
    #be_orders.report()

    to_excel(nl_orders, be_orders, file_name)

# takes the raw data from the CSV and turns it into something useful.. and readable 
# data structure = dictionary with product id as key. Inside is the title and a list of orders
def convert_data(data, nl_data, be_data):
    # only use these rows, ignore the rest
    useful_types = ["Correctie verkoopprijs artikel(en)", "Verkoopprijs artikel(en), ontvangen van kopers en door bol.com door te storten"]

    for row in data:
        if row[0] in useful_types:
            product_id = int(row[2])
            title = row[3]
            total_price = -float(row[12])
            amount = int(row[6])

            country = row[13]

            # if its a return, set amount to negative
            if row[0] == "Correctie verkoopprijs artikel(en)":
                amount *= -1

            if country == "NL":
                nl_data.add_order(product_id, title, total_price, amount)
            elif country == "BE":
                be_data.add_order(product_id, title, total_price, amount)
            else:
                raise Exception("no valid country in provided row")


def to_excel(nl_data, be_data, file_name):
    export_path = f"output/{file_name}.xlsx"

    nl_dataframe = pandas.DataFrame(nl_data.as_dataframe(file_name))
    be_dataframe = pandas.DataFrame(be_data.as_dataframe(file_name))

    #add totals to the bottom of each
    nl_dataframe.loc[len(nl_dataframe)] = ["", "", "", "", "", "Total NL (w/o BTW)", round(nl_data.get_country_total() / 1.21, 2)]
    nl_dataframe.loc[len(nl_dataframe)] = ["", "", "", "", "", "Total NL", nl_data.get_country_total()]
    nl_dataframe.loc[len(nl_dataframe)] = ["", "", "", "", "", "Total NL+BE", nl_data.get_country_total() + be_data.get_country_total()]

    be_dataframe.loc[len(be_dataframe)] = ["", "", "", "", "", "Total BE (w/o BTW)", round(be_data.get_country_total() / 1.21, 2)]
    be_dataframe.loc[len(be_dataframe)] = ["", "", "", "", "", "Total BE ", be_data.get_country_total()]
    be_dataframe.loc[len(be_dataframe)] = ["", "", "", "", "", "Total NL+BE", nl_data.get_country_total() + be_data.get_country_total()]

    print(f"Converting data into {export_path}..")

    with pandas.ExcelWriter(export_path, engine="openpyxl") as writer:
        nl_dataframe.to_excel(writer, sheet_name='NL', index=False)
        be_dataframe.to_excel(writer, sheet_name='BE', index=False)

    print("adjusting column widths..")

    # load into the generated file, get both sheets
    workbook  = load_workbook(export_path)   
    nl_sheet = workbook["NL"] 
    be_sheet = workbook["BE"]

    from openpyxl.styles import numbers # type: ignore

    # for all except the title, the column name is the longest string, so loop over them and set width to column name length
    for letter in ['A', 'C', 'D', 'E', 'F', 'G']:
        nl_sheet.column_dimensions[letter].width = len(nl_sheet[f"{letter}1"].value) + 5
        be_sheet.column_dimensions[letter].width = len(be_sheet[f"{letter}1"].value) + 5

        if letter == "F":
            nl_sheet.column_dimensions[letter].width = len("Total NL (w/o BTW)")
            be_sheet.column_dimensions[letter].width = len("Total BE (w/o BTW)")

    # then, set the width of the title row, B
    max_width = 0
    for row_number in range(1, nl_sheet.max_row + 1):
        cell_value = nl_sheet[f"B{row_number}"].value
        if cell_value is not None:
            if len(cell_value) > max_width:
                max_width = len(cell_value)
    nl_sheet.column_dimensions["B"].width = max_width + 5

    max_width = 0
    for row_number in range(1, be_sheet.max_row + 1):
        cell_value = be_sheet[f"B{row_number}"].value
        if cell_value is not None:
            if len(cell_value) > max_width:
                max_width = len(cell_value)
    be_sheet.column_dimensions["B"].width = max_width + 5

    # Format product ID columns as text
    for sheet in [nl_sheet, be_sheet]:
        for row in range(2, sheet.max_row + 1):  # Skip header row
            cell = sheet['A' + str(row)] 
            cell.number_format = '@'  # '@' is Excel's text format

    workbook.save(export_path)

    print("Done!")

main()